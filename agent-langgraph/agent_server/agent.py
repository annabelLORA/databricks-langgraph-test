import base64
import io
import json
import logging
from datetime import datetime
from typing import AsyncGenerator, Optional

import mlflow
import openpyxl
from databricks.sdk import WorkspaceClient
from databricks_langchain import ChatDatabricks
from langchain.agents import create_agent
from langchain_core.tools import tool
from mlflow.genai.agent_server import invoke, stream
from mlflow.types.responses import (
    ResponsesAgentRequest,
    ResponsesAgentResponse,
    ResponsesAgentStreamEvent,
    to_chat_completions_input,
)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent_server.utils import (
    get_session_id,
    process_agent_astream_events,
)

logger = logging.getLogger(__name__)
mlflow.langchain.autolog()
logging.getLogger("mlflow.utils.autologging_utils").setLevel(logging.ERROR)
sp_workspace_client = WorkspaceClient()

SYSTEM_PROMPT = """You are a construction industry expert assistant for Laing O'Rourke, \
specialising in construction management, engineering processes, safety, quality, and risk.

You handle three types of requests:

**1. General Construction Q&A**
Answer questions about construction methods, materials, engineering principles, \
site management, safety standards, procurement, contracts, and industry best practices. \
Be clear, practical, and accurate.

**2. Process Guidance with References**
When asked about specific construction processes or "what are the next steps for X", \
use the `get_construction_process` tool to retrieve structured step-by-step guidance \
with relevant standards and reference links. Always present steps in a numbered list \
with the reference links clearly shown.

**3. Risk Plan Excel Generation**
When asked to generate a risk plan, risk register, or risk assessment for a project, \
use the `generate_risk_plan_excel` tool. First, analyse the user's project description \
to identify the key risks (likelihood, consequence, controls). Then call the tool with \
those risks. Present the result clearly and tell the user how to decode and save the file.

Always be professional, concise, and safety-conscious. When citing standards, \
prefer AS/NZS standards where applicable (Australian context).
"""

# ── Workflow 2: Process guidance ──────────────────────────────────────────────

CONSTRUCTION_PROCESSES = {
    "excavation": {
        "title": "Excavation & Earthworks",
        "steps": [
            "Obtain site survey and geotechnical investigation report",
            "Submit and approve Excavation Method Statement / Safe Work Method Statement (SWMS)",
            "Identify and mark all underground services (Dial Before You Dig)",
            "Install erosion and sediment controls (silt fences, sediment basins)",
            "Set out excavation boundaries and establish benchmarks",
            "Begin excavation with shoring/benching/battering as per design",
            "Monitor excavation for groundwater, instability, and contamination",
            "Complete inspection and approval prior to blinding/formwork",
        ],
        "references": [
            ("AS 3798 – Guidelines on earthworks", "https://www.standards.org.au/standards-catalogue/sa-snz/building/me-001/as-3798-2007"),
            ("Safe Work Australia – Excavation SWMS", "https://www.safeworkaustralia.gov.au/doc/model-code-of-practice-excavation-work"),
            ("Dial Before You Dig", "https://www.1100.com.au"),
        ],
    },
    "concrete": {
        "title": "Concrete Placement",
        "steps": [
            "Review structural drawings and mix design specifications",
            "Inspect and approve formwork, reinforcement, and embedments",
            "Conduct pre-pour inspection and obtain approval",
            "Order concrete to approved mix design; verify delivery dockets",
            "Place and compact concrete (vibrate to eliminate voids)",
            "Finish surface to specified tolerances",
            "Implement curing regime (wet hessian, curing compound, or formwork retention)",
            "Test cubes/cylinders taken per AS 1012; record results",
            "Strip formwork at approved time and inspect for defects",
        ],
        "references": [
            ("AS 3600 – Concrete Structures", "https://www.standards.org.au/standards-catalogue/sa-snz/building/bd-002/as-3600-2018"),
            ("AS 1379 – Specification and Supply of Concrete", "https://www.standards.org.au/standards-catalogue/sa-snz/building/bd-001/as-1379-2007"),
            ("AS 1012 – Methods of Testing Concrete", "https://www.standards.org.au/standards-catalogue/sa-snz/building/bd-001/as-1012-9-2014"),
        ],
    },
    "scaffolding": {
        "title": "Scaffolding Erection & Use",
        "steps": [
            "Determine scaffold type (system, tube-and-coupler, suspended) per task requirements",
            "Engage licensed scaffolder (Class 2 or above as required)",
            "Prepare scaffold design and load calculations for complex structures",
            "Inspect base plates, sole boards, and ground bearing capacity",
            "Erect scaffold with ties, bracing, and guardrails per design",
            "Conduct handover inspection and issue Scafftag / scaffold tag",
            "Conduct weekly inspections and after any adverse weather event",
            "Ensure all users complete scaffold awareness training",
            "Dismantle scaffold in reverse order under scaffolder supervision",
        ],
        "references": [
            ("AS/NZS 4576 – Guidelines for scaffolding", "https://www.standards.org.au/standards-catalogue/sa-snz/building/me-001/as-nzs-4576-1995"),
            ("Safe Work Australia – Scaffolding Code", "https://www.safeworkaustralia.gov.au/doc/model-code-of-practice-managing-risks-falls-general-construction"),
            ("WorkSafe Scaffold Licensing", "https://www.safeworkaustralia.gov.au/licensing"),
        ],
    },
    "commissioning": {
        "title": "Commissioning & Handover",
        "steps": [
            "Develop Commissioning Management Plan aligned to contract requirements",
            "Complete pre-commissioning punch list and close-out all items",
            "Conduct factory acceptance tests (FAT) and site acceptance tests (SAT)",
            "Verify all O&M manuals, warranties, and as-built drawings are complete",
            "Complete integrated systems testing (IST) across disciplines",
            "Conduct training for client's operations and maintenance staff",
            "Obtain all regulatory approvals, certificates of occupancy/completion",
            "Achieve Practical Completion milestone and issue to client",
            "Monitor defects liability period (DLP) and close-out defects",
        ],
        "references": [
            ("AIPM – Commissioning Guidance", "https://www.aipm.com.au"),
            ("ISO 10005 – Quality Management Plans", "https://www.iso.org/standard/72621.html"),
            ("Laing O'Rourke Project Controls Framework", "https://www.laingorourke.com"),
        ],
    },
    "default": {
        "title": "General Construction Process",
        "steps": [
            "Confirm scope, specifications, and design documentation are approved",
            "Prepare and approve Method Statement / Safe Work Method Statement",
            "Check permits, approvals, and hold points required",
            "Confirm resources (labour, plant, materials) are mobilised and inspected",
            "Conduct pre-work toolbox talk and site induction for all workers",
            "Execute works to approved method statement with ITP sign-offs",
            "Conduct inspections at each hold point; obtain approvals before proceeding",
            "Document quality records, non-conformances, and corrective actions",
            "Complete works and obtain sign-off for next phase or handover",
        ],
        "references": [
            ("AS/NZS ISO 45001 – Occupational Health and Safety", "https://www.standards.org.au/standards-catalogue/sa-snz/publicsafety/sf-001/as-nzs-iso-45001-2018"),
            ("AS/NZS ISO 9001 – Quality Management Systems", "https://www.standards.org.au/standards-catalogue/sa-snz/generaltechnologies/qr-001/as-nzs-iso-9001-2016"),
            ("Safe Work Australia", "https://www.safeworkaustralia.gov.au"),
        ],
    },
}


@tool
def get_construction_process(process_name: str) -> str:
    """Get step-by-step guidance and reference links for a named construction process.

    Args:
        process_name: The construction process to look up (e.g. 'excavation', 'concrete',
                      'scaffolding', 'commissioning', or any general description).
    """
    key = process_name.lower().strip()
    # fuzzy match on keywords
    matched = next(
        (v for k, v in CONSTRUCTION_PROCESSES.items() if k != "default" and k in key),
        CONSTRUCTION_PROCESSES["default"],
    )
    steps_text = "\n".join(f"  {i+1}. {s}" for i, s in enumerate(matched["steps"]))
    refs_text = "\n".join(f"  - [{name}]({url})" for name, url in matched["references"])
    return (
        f"## {matched['title']}\n\n"
        f"### Steps\n{steps_text}\n\n"
        f"### Reference Standards & Links\n{refs_text}"
    )


# ── Workflow 3: Risk plan Excel generation ────────────────────────────────────

RISK_COLORS = {
    "Extreme": "C00000",
    "High": "FF0000",
    "Medium": "FFC000",
    "Low": "92D050",
}

LIKELIHOOD_SCORE = {"Rare": 1, "Unlikely": 2, "Possible": 3, "Likely": 4, "Almost Certain": 5}
CONSEQUENCE_SCORE = {"Insignificant": 1, "Minor": 2, "Moderate": 3, "Major": 4, "Catastrophic": 5}


def _risk_rating(likelihood: str, consequence: str) -> str:
    l = LIKELIHOOD_SCORE.get(likelihood, 3)
    c = CONSEQUENCE_SCORE.get(consequence, 3)
    score = l * c
    if score >= 15:
        return "Extreme"
    elif score >= 8:
        return "High"
    elif score >= 4:
        return "Medium"
    return "Low"


def _build_excel(project_name: str, risks: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Risk Plan — {project_name}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y')}    |    Organisation: Laing O'Rourke"
    ws["A2"].font = Font(italic=True, color="595959")
    ws.merge_cells("A2:J2")

    # Headers
    headers = [
        "Risk ID", "Risk Description", "Category", "Likelihood",
        "Consequence", "Risk Rating", "Controls / Mitigation",
        "Residual Likelihood", "Residual Consequence", "Residual Rating",
    ]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 36

    # Data rows
    for i, risk in enumerate(risks, 1):
        row = header_row + i
        likelihood = risk.get("likelihood", "Possible")
        consequence = risk.get("consequence", "Moderate")
        residual_likelihood = risk.get("residual_likelihood", "Unlikely")
        residual_consequence = risk.get("residual_consequence", "Minor")
        rating = _risk_rating(likelihood, consequence)
        residual_rating = _risk_rating(residual_likelihood, residual_consequence)

        values = [
            f"R{i:02d}",
            risk.get("description", ""),
            risk.get("category", "General"),
            likelihood,
            consequence,
            rating,
            risk.get("controls", ""),
            residual_likelihood,
            residual_consequence,
            residual_rating,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col in (6, 10):  # rating columns
                color = RISK_COLORS.get(val, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF" if val in ("Extreme", "High") else "000000")
        ws.row_dimensions[row].height = 45

    # Column widths
    col_widths = [8, 40, 18, 16, 16, 14, 50, 18, 18, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Risk matrix legend sheet
    ws2 = wb.create_sheet("Risk Matrix")
    ws2["A1"] = "Risk Rating Matrix"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.merge_cells("A1:F1")

    matrix_headers = ["", "Insignificant", "Minor", "Moderate", "Major", "Catastrophic"]
    likelihood_labels = ["Almost Certain", "Likely", "Possible", "Unlikely", "Rare"]
    matrix_colors = [
        ["FF0000", "FF0000", "C00000", "C00000", "C00000"],
        ["FFC000", "FF0000", "FF0000", "C00000", "C00000"],
        ["92D050", "FFC000", "FF0000", "FF0000", "C00000"],
        ["92D050", "92D050", "FFC000", "FFC000", "FF0000"],
        ["92D050", "92D050", "92D050", "FFC000", "FFC000"],
    ]

    for col, h in enumerate(matrix_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r, label in enumerate(likelihood_labels, 1):
        ws2.cell(row=3 + r, column=1, value=label).font = Font(bold=True)
        for c, color in enumerate(matrix_colors[r - 1], 2):
            cell = ws2.cell(row=3 + r, column=c)
            cell.fill = PatternFill("solid", fgColor=color)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@tool
def generate_risk_plan_excel(project_name: str, risks_json: str) -> str:
    """Generate a risk plan Excel workbook for a construction project.

    Args:
        project_name: Name of the project.
        risks_json: JSON array of risk objects. Each object should have:
            - description (str): What the risk is
            - category (str): e.g. 'Safety', 'Programme', 'Cost', 'Environment', 'Quality'
            - likelihood (str): 'Rare' | 'Unlikely' | 'Possible' | 'Likely' | 'Almost Certain'
            - consequence (str): 'Insignificant' | 'Minor' | 'Moderate' | 'Major' | 'Catastrophic'
            - controls (str): Mitigation measures
            - residual_likelihood (str): Same scale as likelihood after controls
            - residual_consequence (str): Same scale as consequence after controls

    Returns:
        JSON with base64 encoded Excel file content and filename.
    """
    try:
        risks = json.loads(risks_json)
    except json.JSONDecodeError as e:
        return f"Error parsing risks JSON: {e}"

    excel_bytes = _build_excel(project_name, risks)
    b64 = base64.b64encode(excel_bytes).decode("utf-8")
    filename = f"risk_plan_{project_name.replace(' ', '_').lower()}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return json.dumps({
        "filename": filename,
        "content_base64": b64,
        "risk_count": len(risks),
        "message": (
            f"Excel risk plan generated with {len(risks)} risks. "
            f"Filename: {filename}. "
            "To save: copy the content_base64 value, decode from base64, and save as .xlsx. "
            "In Python: import base64; open('risk_plan.xlsx','wb').write(base64.b64decode('<content_base64>'))"
        ),
    })


# ── Agent initialisation ──────────────────────────────────────────────────────

@tool
def get_current_time() -> str:
    """Get the current date and time."""
    return datetime.now().isoformat()


async def init_agent(workspace_client=None):
    tools = [get_current_time, get_construction_process, generate_risk_plan_excel]
    return create_agent(
        tools=tools,
        model=ChatDatabricks(endpoint="databricks-gpt-5-2"),
        state_modifier=SYSTEM_PROMPT,
    )


@invoke()
async def invoke_handler(request: ResponsesAgentRequest) -> ResponsesAgentResponse:
    outputs = [
        event.item
        async for event in stream_handler(request)
        if event.type == "response.output_item.done"
    ]
    return ResponsesAgentResponse(output=outputs)


@stream()
async def stream_handler(
    request: ResponsesAgentRequest,
) -> AsyncGenerator[ResponsesAgentStreamEvent, None]:
    if session_id := get_session_id(request):
        mlflow.update_current_trace(metadata={"mlflow.trace.session": session_id})

    agent = await init_agent()
    messages = {"messages": to_chat_completions_input([i.model_dump() for i in request.input])}

    async for event in process_agent_astream_events(
        agent.astream(input=messages, stream_mode=["updates", "messages"])
    ):
        yield event
