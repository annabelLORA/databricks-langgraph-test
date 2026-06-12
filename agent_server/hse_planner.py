"""HSE Risk Planning: category classification, control selection, and Excel generation."""
import copy
import io
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from agent_server.knowledge import DATA_DIR, HSE_CONTROLS, get_control

# ── Category mapping ───────────────────────────────────────────────────────────

VALID_CATEGORIES = [
    "Asbestos",
    "Barricading",
    "Confined Space",
    "Cranes and Lifting",
    "Demolition",
    "Drill and Blast Management",
    "Excavation & Ground Penetration",
    "Fitness for Work",
    "Formwork & Falsework",
    "General Electrical Safety",
    "Hazardous Energy",
    "Hazardous Substances",
    "Heavy Vehicles",
    "Logistics",
    "Manual Handling",
    "Permit to Work",
    "Personal Protective Equipment",
    "Piling",
    "Plant and Equipment",
    "Precast Concrete",
    "Prevention of Falls & Dropped Objects",
    "Psychosocial Hazards",
    "Respirable Crystalline Silica",
    "Scaffolding",
    "Site Establishment",
    "Traffic Management",
    "Working in Rail Environments",
    "Work In, Over or Adjacent to Water",
    "Air Quality and Dust Management",
    "Environmental Boundary Delineation",
    "Biodiversity and Biosecurity",
    "Erosion and Sediment Control",
    "Dangerous Goods and Chemical Management",
    "Heritage",
    "Noise and Vibration",
    "Soil Management",
    "Spoil, Waste, Resource and Recovery",
    "Temporary Waterway Structures",
    "Water Management",
]

# Checklist name → canonical category name
CHECKLIST_TO_CATEGORY: dict[str, str] = {
    "FSR Excavation": "Excavation & Ground Penetration",
    "FSR Preventions of Falls & Dropped Objects": "Prevention of Falls & Dropped Objects",
    "FSR Work On, Over & Adjacent to Water": "Work In, Over or Adjacent to Water",
    "FSR Work in Rail Environments": "Working in Rail Environments",
    "FSR Hazardous Dust & Fibres": "Respirable Crystalline Silica",
    "FSR Cranes & Lifting": "Cranes and Lifting",
    "FSR Confined Spaces": "Confined Space",
    "FSR Demolition": "Demolition",
    "FSR Formwork & Falsework": "Formwork & Falsework",
    "FSR General Electrical Safety": "General Electrical Safety",
    "FSR Hazardous Energy": "Hazardous Energy",
    "FSR Heavy Vehicles": "Heavy Vehicles",
    "FSR Piling": "Piling",
    "FSR Plant and Equipment": "Plant and Equipment",
    "FSR Precast Concrete": "Precast Concrete",
    "FSR Scaffolding": "Scaffolding",
    "FSR Traffic Management": "Traffic Management",
    "SER Air Quality & Dust": "Air Quality and Dust Management",
    "SER Biodiversity": "Biodiversity and Biosecurity",
    "SER Biosecurity": "Biodiversity and Biosecurity",
    "SER Cultural and European Heritage": "Heritage",
    "SER Surface Water Management": "Water Management",
    "SER Groundwater Management": "Water Management",
    "SER Groundwater Dewatering": "Water Management",
    "SER Spoil and Waste Management": "Spoil, Waste, Resource and Recovery",
    "SER Contamination Management": "Hazardous Substances",
    "SER Dangerous Goods and Chemical Management": "Dangerous Goods and Chemical Management",
    "SER Erosion and Sediment Control": "Erosion and Sediment Control",
    "SER Noise and Vibration": "Noise and Vibration",
    "SER Temporary Waterway Structures": "Temporary Waterway Structures",
    "SER ERSED Inspection": "Erosion and Sediment Control",
}

# Keyword patterns → risk categories (order: most specific first)
# Each entry: (keyword, [categories])
KEYWORD_CATEGORIES: list[tuple[str, list[str]]] = [
    ("hdd", ["Excavation & Ground Penetration"]),
    ("horizontal directional drill", ["Excavation & Ground Penetration", "Respirable Crystalline Silica"]),
    ("pile", ["Piling", "Work In, Over or Adjacent to Water"]),
    ("piling", ["Piling", "Work In, Over or Adjacent to Water"]),
    ("demolit", ["Demolition", "Prevention of Falls & Dropped Objects"]),
    ("remove redundant", ["Excavation & Ground Penetration", "Hazardous Energy"]),
    ("scaffold", ["Scaffolding", "Prevention of Falls & Dropped Objects"]),
    ("crane", ["Cranes and Lifting", "Prevention of Falls & Dropped Objects"]),
    ("lift", ["Cranes and Lifting"]),
    ("hoist", ["Cranes and Lifting"]),
    ("formwork", ["Formwork & Falsework", "Prevention of Falls & Dropped Objects"]),
    ("falsework", ["Formwork & Falsework"]),
    ("concrete", ["Formwork & Falsework", "Respirable Crystalline Silica"]),
    ("pour", ["Formwork & Falsework"]),
    ("curing", ["Formwork & Falsework"]),
    ("reinforcement", ["Formwork & Falsework", "Prevention of Falls & Dropped Objects"]),
    ("reo", ["Formwork & Falsework", "Prevention of Falls & Dropped Objects"]),
    ("precast", ["Precast Concrete", "Cranes and Lifting"]),
    ("permadec", ["Precast Concrete", "Cranes and Lifting"]),
    ("segment", ["Precast Concrete", "Cranes and Lifting"]),
    ("beam", ["Cranes and Lifting", "Prevention of Falls & Dropped Objects"]),
    ("girder", ["Cranes and Lifting", "Prevention of Falls & Dropped Objects"]),
    ("stay cable", ["Prevention of Falls & Dropped Objects", "Cranes and Lifting"]),
    ("tower", ["Prevention of Falls & Dropped Objects", "Scaffolding"]),
    ("parapet", ["Prevention of Falls & Dropped Objects"]),
    ("barrier", ["Traffic Management", "Prevention of Falls & Dropped Objects"]),
    ("waterproof", ["Hazardous Substances", "Prevention of Falls & Dropped Objects"]),
    ("electrical", ["General Electrical Safety"]),
    ("electri", ["General Electrical Safety"]),
    ("light", ["General Electrical Safety"]),
    ("its", ["General Electrical Safety"]),
    ("commission", ["General Electrical Safety", "Plant and Equipment"]),
    ("walkdown", ["Plant and Equipment", "Prevention of Falls & Dropped Objects"]),
    ("road safety audit", ["Traffic Management"]),
    ("traffic", ["Traffic Management"]),
    ("road", ["Traffic Management", "Plant and Equipment"]),
    ("subgrade", ["Excavation & Ground Penetration", "Plant and Equipment"]),
    ("subbase", ["Excavation & Ground Penetration", "Plant and Equipment"]),
    ("excavat", ["Excavation & Ground Penetration"]),
    ("trench", ["Excavation & Ground Penetration"]),
    ("backfill", ["Excavation & Ground Penetration", "Plant and Equipment"]),
    ("earthwork", ["Excavation & Ground Penetration", "Plant and Equipment"]),
    ("jetty", ["Work In, Over or Adjacent to Water", "Temporary Waterway Structures"]),
    ("marine", ["Work In, Over or Adjacent to Water"]),
    ("water", ["Work In, Over or Adjacent to Water"]),
    ("river", ["Work In, Over or Adjacent to Water", "Temporary Waterway Structures"]),
    ("pontoon", ["Work In, Over or Adjacent to Water"]),
    ("scour", ["Work In, Over or Adjacent to Water", "Temporary Waterway Structures"]),
    ("deck", ["Prevention of Falls & Dropped Objects", "Formwork & Falsework"]),
    ("frp", ["Prevention of Falls & Dropped Objects", "Hazardous Substances"]),
    ("paint", ["Hazardous Substances", "Prevention of Falls & Dropped Objects"]),
    ("graffiti", ["Hazardous Substances"]),
    ("primer", ["Hazardous Substances"]),
    ("silica", ["Respirable Crystalline Silica"]),
    ("dust", ["Respirable Crystalline Silica", "Air Quality and Dust Management"]),
    ("drilling", ["Excavation & Ground Penetration", "Respirable Crystalline Silica"]),
    ("grind", ["Respirable Crystalline Silica"]),
    ("cut", ["Respirable Crystalline Silica", "Noise and Vibration"]),
    ("noise", ["Noise and Vibration"]),
    ("vibrat", ["Noise and Vibration"]),
    ("vehicle", ["Heavy Vehicles", "Traffic Management"]),
    ("truck", ["Heavy Vehicles"]),
    ("haulage", ["Heavy Vehicles"]),
    ("delivery", ["Heavy Vehicles", "Logistics"]),
    ("mobilisation", ["Plant and Equipment", "Heavy Vehicles"]),
    ("plant", ["Plant and Equipment"]),
    ("equipment", ["Plant and Equipment"]),
    ("machin", ["Plant and Equipment"]),
    ("install", ["Plant and Equipment", "Prevention of Falls & Dropped Objects"]),
    ("erect", ["Prevention of Falls & Dropped Objects"]),
    ("height", ["Prevention of Falls & Dropped Objects"]),
    ("fall", ["Prevention of Falls & Dropped Objects"]),
    ("edge protection", ["Prevention of Falls & Dropped Objects"]),
    ("confined", ["Confined Space"]),
    ("asbestos", ["Asbestos"]),
    ("hazardous", ["Hazardous Substances"]),
    ("chemical", ["Dangerous Goods and Chemical Management"]),
    ("fuel", ["Dangerous Goods and Chemical Management"]),
    ("heritage", ["Heritage"]),
    ("cultural", ["Heritage"]),
    ("biodiversity", ["Biodiversity and Biosecurity"]),
    ("ecological", ["Biodiversity and Biosecurity"]),
    ("erosion", ["Erosion and Sediment Control"]),
    ("sediment", ["Erosion and Sediment Control"]),
    ("ersed", ["Erosion and Sediment Control"]),
    ("groundwater", ["Water Management"]),
    ("dewater", ["Water Management"]),
    ("surface water", ["Water Management"]),
    ("waterway", ["Temporary Waterway Structures"]),
    ("rail", ["Working in Rail Environments"]),
    ("establish", ["Site Establishment"]),
    ("compound", ["Site Establishment"]),
    ("spoil", ["Spoil, Waste, Resource and Recovery"]),
    ("waste", ["Spoil, Waste, Resource and Recovery"]),
    ("lockout", ["Hazardous Energy"]),
    ("isolation", ["Hazardous Energy"]),
    ("permit", ["Permit to Work"]),
    ("confined space", ["Confined Space"]),
    ("punchlist", ["Plant and Equipment", "General Electrical Safety"]),
]

# Cross-cutting controls always applied (checklist names)
CROSS_CUTTING_ALWAYS = [
    "Pre-Start Toolbox Talks",
    "Critical Risk Briefings",
    "AU SWMS Part C – Task Observation",
]

# FSR checklist → category (for primary FSR selection)
FSR_CHECKLIST_TO_CATEGORY = {k: v for k, v in CHECKLIST_TO_CATEGORY.items() if k.startswith("FSR")}
SER_CHECKLIST_TO_CATEGORY = {k: v for k, v in CHECKLIST_TO_CATEGORY.items() if k.startswith("SER")}

# Category → FSR checklist (reverse map; only categories with a direct FSR)
CATEGORY_TO_FSR: dict[str, str] = {v: k for k, v in FSR_CHECKLIST_TO_CATEGORY.items()}
CATEGORY_TO_SER: dict[str, str] = {}
for checklist, cat in SER_CHECKLIST_TO_CATEGORY.items():
    if cat not in CATEGORY_TO_SER:
        CATEGORY_TO_SER[cat] = checklist

# ── Classification ─────────────────────────────────────────────────────────────


def classify_activity(description: str, wbs: str = "") -> list[str]:
    """Return relevant risk categories for an activity, most important first."""
    text = (description + " " + wbs).lower()
    found: list[str] = []
    seen: set[str] = set()
    for kw, cats in KEYWORD_CATEGORIES:
        if kw in text:
            for c in cats:
                if c not in seen:
                    seen.add(c)
                    found.append(c)
    if not found:
        found = ["Plant and Equipment", "Prevention of Falls & Dropped Objects"]
    return found[:6]


def primary_category(categories: list[str]) -> str:
    """Return the primary category — the first one with a direct FSR."""
    for cat in categories:
        if cat in CATEGORY_TO_FSR:
            return cat
    return categories[0]


# ── Control selection ──────────────────────────────────────────────────────────


def controls_for_task(categories: list[str], primary: str, description: str) -> list[dict]:
    """Return the ordered list of controls for a task.

    Each control: {checklist, module, type, category}
    """
    selected: list[dict] = []
    seen_titles: set[str] = set()
    desc_lower = description.lower()

    def add(checklist: str, category: str) -> None:
        if checklist in seen_titles:
            return
        ctrl = get_control(checklist)
        if ctrl:
            seen_titles.add(checklist)
            selected.append(
                {
                    "checklist": checklist,
                    "module": ctrl["module"],
                    "type": ctrl["type"],
                    "category": category,
                }
            )

    # 1. FSR/SER for each applicable category
    for cat in categories:
        fsr = CATEGORY_TO_FSR.get(cat)
        if fsr:
            add(fsr, cat)
        ser = CATEGORY_TO_SER.get(cat)
        if ser:
            add(ser, cat)

    # 2. Task-specific conditional controls
    if "concrete boom pump" in desc_lower or "boom pump" in desc_lower:
        add("AU Concrete Boom Pump Setup Checklist", primary)
    if any(k in desc_lower for k in ("deliver", "haulage", "heavy vehicle", "truck")):
        add("AU CoR - Load Restraint Inspection", primary)
        add("AU CoR - Fatigue Management Inspection", primary)
    if any(k in desc_lower for k in ("water", "jetty", "marine", "river")):
        add("AU ERSED Inspection", "Erosion and Sediment Control")
    if any(k in desc_lower for k in ("demolit", "removal")):
        add("AU SWMS Part C – Task Observation", primary)

    # 3. Cross-cutting always-on
    for cc in CROSS_CUTTING_ALWAYS:
        add(cc, primary)

    return selected


# ── Risk text generation ───────────────────────────────────────────────────────

_RISK_DETAILS: dict[str, str] = {
    "Cranes and Lifting": (
        "During {task}, crane and rigging operations expose workers to uncontrolled load movement, "
        "rigging failure, or crane overload. Without this control, personnel risk being struck by a "
        "falling or swinging load, causing fatal or serious crush injuries."
    ),
    "Prevention of Falls & Dropped Objects": (
        "Personnel undertaking {task} are exposed to unprotected edges, openings, or elevated work "
        "platforms. Dropped tools, materials, or components from height could strike workers or "
        "members of the public below, resulting in fatal or serious injury."
    ),
    "Formwork & Falsework": (
        "Formwork and falsework associated with {task} present risks of premature collapse under "
        "wet concrete loads or inadequate propping. Failure exposes workers to being engulfed by "
        "concrete, structural collapse, and serious crush or burial injuries."
    ),
    "Work In, Over or Adjacent to Water": (
        "Works comprising {task} place personnel in proximity to open water, creating immersion and "
        "drowning risk. Adverse weather, tidal or current changes, and falling from structures over "
        "water may result in rapid incapacitation and fatality without appropriate controls."
    ),
    "Excavation & Ground Penetration": (
        "Ground penetration activities forming part of {task} expose workers to soil instability, "
        "uncontrolled collapse, and the risk of striking pressurised or live underground services. "
        "Inadequate controls may result in burial, electrocution, or explosion."
    ),
    "Scaffolding": (
        "Scaffold installation, modification, or dismantling during {task} exposes workers to falls "
        "from height and structural failure. Competency gaps and inadequate inspections increase the "
        "likelihood of collapse under load, resulting in fatal or serious injury."
    ),
    "Piling": (
        "Piling operations associated with {task} generate high-energy dynamic loads, noise, and "
        "vibration. Workers risk being struck by moving pile equipment, and ground movement may "
        "destabilise adjacent structures or expose utilities."
    ),
    "Traffic Management": (
        "Interaction of construction vehicles and pedestrians during {task} in an active traffic "
        "environment presents a vehicle strike risk. Inadequate traffic management controls may "
        "result in a worker being struck by a moving vehicle, causing fatal or serious injury."
    ),
    "Demolition": (
        "Demolition activities forming part of {task} expose workers to uncontrolled structural "
        "collapse, falling debris, and release of hazardous materials. Without this control, workers "
        "may suffer fatal or serious injury from unplanned collapse or impact from falling elements."
    ),
    "Precast Concrete": (
        "Precast concrete handling during {task} involves heavy element lifts under tension. "
        "Rigging failure, improper pick-point attachment, or uncontrolled element swing can cause "
        "elements to fall or collide with personnel, resulting in crush or fatal injury."
    ),
    "General Electrical Safety": (
        "Electrical installation and commissioning activities associated with {task} expose workers "
        "to live conductors, inadvertent energisation, and arc flash hazards. Without this control, "
        "electrocution or electrical burns represent a significant risk of fatality."
    ),
    "Hazardous Energy": (
        "Energy isolation requirements during {task} address stored mechanical, hydraulic, or "
        "electrical energy. Failure to implement effective lockout/tagout procedures may result in "
        "unexpected equipment energisation, causing severe crush, entanglement, or electrocution."
    ),
    "Hazardous Substances": (
        "Hazardous substances used or disturbed during {task} expose workers to chemical burns, "
        "inhalation of toxic vapours, or skin sensitisation. Inadequate controls may result in "
        "acute poisoning, long-term occupational disease, or environmental contamination."
    ),
    "Respirable Crystalline Silica": (
        "Cutting, grinding, or drilling activities associated with {task} generate respirable "
        "crystalline silica dust. Chronic inhalation without adequate controls leads to silicosis, "
        "an irreversible and potentially fatal lung disease."
    ),
    "Plant and Equipment": (
        "Mobile plant and powered equipment operating during {task} create exclusion zone risks, "
        "operator visibility blind spots, and mechanical failure hazards. Personnel in proximity to "
        "operating plant risk being struck, run over, or crushed."
    ),
    "Heavy Vehicles": (
        "Heavy vehicle movements associated with {task} in a construction environment create "
        "vehicle-pedestrian conflict risks. Drivers with restricted visibility and workers on foot "
        "sharing the same corridor increases the likelihood of a fatal vehicle strike."
    ),
    "Noise and Vibration": (
        "Construction activities during {task} generate sustained noise and ground-borne vibration "
        "levels exceeding occupational exposure limits. Without controls, workers face noise-induced "
        "hearing loss and hand-arm vibration syndrome as long-term health consequences."
    ),
    "Erosion and Sediment Control": (
        "Ground disturbance and exposed surfaces during {task} present a risk of erosion and "
        "uncontrolled sediment runoff into sensitive waterways. Ineffective controls may result in "
        "regulatory non-compliance, environmental harm, and community impact."
    ),
    "Water Management": (
        "Dewatering and groundwater management associated with {task} risk drawing contaminated "
        "groundwater to surface or discharging to sensitive receiving environments without treatment. "
        "Inadequate management may cause environmental harm and regulatory breach."
    ),
    "Temporary Waterway Structures": (
        "Temporary in-water structures forming part of {task} must resist hydraulic loads and "
        "maintain environmental containment. Failure of cofferdams or temporary works may result in "
        "catastrophic inundation, drowning, and uncontrolled sediment release to waterways."
    ),
    "Dangerous Goods and Chemical Management": (
        "Storage and handling of dangerous goods during {task} presents a spill, fire, and "
        "inhalation risk to workers and the environment. Inadequate containment and emergency "
        "response may escalate a minor spill to a major environmental incident or explosion."
    ),
    "Air Quality and Dust Management": (
        "Earthworks, demolition, and material handling during {task} generate airborne particulates "
        "affecting workers and neighbouring receptors. Sustained dust emissions without controls "
        "risk respiratory health impacts and community and regulatory complaints."
    ),
    "Biodiversity and Biosecurity": (
        "Construction activities during {task} in proximity to sensitive ecological areas risk "
        "disturbance to protected flora, fauna, and habitat. Biosecurity failures may introduce "
        "invasive species, resulting in regulatory breach and irreversible ecological harm."
    ),
    "Heritage": (
        "Ground disturbance and construction impacts during {task} in areas of cultural or heritage "
        "significance risk undiscovered artefacts being damaged or destroyed. Failure to manage this "
        "risk may result in regulatory breach and irreversible loss of heritage value."
    ),
    "Spoil, Waste, Resource and Recovery": (
        "Excavated spoil and construction waste generated during {task} must be classified, "
        "segregated, and disposed of to approved facilities. Improper management risks contamination "
        "of land and water and exposes the project to regulatory non-compliance."
    ),
    "Confined Space": (
        "Confined space entry required during {task} exposes workers to oxygen deficiency, toxic "
        "atmospheric contaminants, or engulfment hazards. Without rigorous entry controls, rescue "
        "capability, and atmospheric monitoring, fatalities are likely."
    ),
    "Site Establishment": (
        "Site establishment activities during {task} involve interface of multiple trades, mobile "
        "plant, and utility connections in an evolving work environment. Without systematic controls, "
        "workers risk being struck by plant, contacting live services, or falling."
    ),
    "Logistics": (
        "Logistics and materials handling operations during {task} create vehicle-pedestrian "
        "interface risks and manual handling exposures. Without controls, workers risk being struck "
        "by delivery vehicles or suffering musculoskeletal injury from unsupported loads."
    ),
}

_DEFAULT_RISK_DETAIL = (
    "Personnel undertaking {task} are exposed to the hazards associated with this work category. "
    "Without implementation of this control, the likelihood and consequence of an incident "
    "resulting in injury, illness, or environmental harm is materially increased."
)

_REASONING: dict[str, str] = {
    "FSR Cranes & Lifting": (
        "This FSR inspection verifies crane documentation, operator competency, lift plans, "
        "exclusion zones, and rigging integrity prior to and during all lifting operations for "
        "{task}. {top5}"
    ),
    "FSR Preventions of Falls & Dropped Objects": (
        "This FSR inspection confirms edge protection, working-at-height permits, harness "
        "inspection, and dropped-object prevention systems are in place for {task}. {top5}"
    ),
    "FSR Work On, Over & Adjacent to Water": (
        "This FSR inspection verifies life rings, rescue boats, spotter arrangements, and personal "
        "flotation device compliance for all works at or over water during {task}. {top5}"
    ),
    "FSR Excavation": (
        "This FSR inspection confirms Dial Before You Dig clearances, shoring/battering compliance, "
        "exclusion zones, and atmospheric monitoring are in place for ground-penetration activities "
        "during {task}. {top5}"
    ),
    "FSR Scaffolding": (
        "This FSR inspection verifies scaffold design sign-off, handover tag status, tie and brace "
        "integrity, and licensed scaffolder supervision for all scaffold use during {task}. {top5}"
    ),
    "FSR Formwork & Falsework": (
        "This FSR inspection confirms falsework propping loads, pour rate controls, form-tie "
        "integrity, and independent design certification are in place before each concrete pour "
        "associated with {task}. {top5}"
    ),
    "FSR Piling": (
        "This FSR inspection verifies pile equipment set-up, exclusion zones, operator competency, "
        "and proximity to utilities and structures prior to piling operations during {task}. {top5}"
    ),
    "FSR Traffic Management": (
        "This FSR inspection confirms TMP implementation, delineation device placement, speed zone "
        "compliance, and flagperson competency for all traffic interfaces during {task}. {top5}"
    ),
    "FSR Demolition": (
        "This FSR inspection verifies structural engineer sign-off, safe demolition sequence, "
        "exclusion zones, and hazardous material surveys prior to all demolition activity during "
        "{task}. {top5}"
    ),
    "FSR Precast Concrete": (
        "This FSR inspection confirms lifting inserts, rigging documentation, element survey, and "
        "structural engineer approval before each precast lift associated with {task}. {top5}"
    ),
    "FSR General Electrical Safety": (
        "This FSR inspection verifies electrical isolation, licensed electrician attendance, "
        "test-and-tag compliance, and live work permits for all electrical activities during "
        "{task}. {top5}"
    ),
    "FSR Hazardous Energy": (
        "This FSR inspection confirms isolation registers, lockout/tagout procedures, verification "
        "of zero energy state, and competency of isolating personnel for {task}. {top5}"
    ),
    "FSR Plant and Equipment": (
        "This FSR inspection verifies plant pre-start checks, operator licences and competency, "
        "exclusion zone implementation, and plant fitness-for-service for all equipment used "
        "during {task}. {top5}"
    ),
    "FSR Heavy Vehicles": (
        "This FSR inspection confirms vehicle roadworthiness, load restraint compliance, driver "
        "fatigue management, and traffic management interface controls for heavy vehicle movements "
        "associated with {task}. {top5}"
    ),
    "FSR Hazardous Dust & Fibres": (
        "This FSR inspection verifies dust suppression measures, RPE fit-testing and issue, "
        "airborne monitoring, and health surveillance arrangements for silica-generating work "
        "during {task}. {top5}"
    ),
    "FSR Confined Spaces": (
        "This FSR inspection confirms entry permits, atmospheric testing, rescue arrangements, "
        "and attendant-worker protocols are in place before each confined space entry during "
        "{task}. {top5}"
    ),
    "FSR Work in Rail Environments": (
        "This FSR inspection verifies track access authorities, protection officer competency, "
        "protection arrangements, and lookout provisions for all rail interface work during "
        "{task}. {top5}"
    ),
    "Pre-Start Toolbox Talks": (
        "The Pre-Start Toolbox Talk ensures all workers on {task} are briefed on daily hazards, "
        "controls, permit conditions, and emergency arrangements before work commences. "
        "A communication gap at task start elevates the probability of an uncontrolled incident. {top5}"
    ),
    "Critical Risk Briefings": (
        "Critical Risk Briefings ensure workers on {task} have formal understanding of LOR's "
        "critical risk standards applicable to the work. Inadequate competency in critical risk "
        "recognition is a root cause in major incident analyses. {top5}"
    ),
    "AU SWMS Part C – Task Observation": (
        "The SWMS Part C Task Observation verifies that actual work practices for {task} match "
        "the approved SWMS. It identifies field-level deviations before they escalate to incidents "
        "and drives corrective actions in real time. {top5}"
    ),
    "AU Concrete Boom Pump Setup Checklist": (
        "This checklist verifies boom pump outrigger setup, ground bearing capacity, proximity to "
        "overhead services, and operator licence for concrete placements during {task}. {top5}"
    ),
    "AU CoR - Load Restraint Inspection": (
        "The CoR Load Restraint Inspection confirms vehicle loading meets Chain of Responsibility "
        "obligations for heavy vehicles associated with {task}, reducing the risk of load shift "
        "causing a road traffic incident. {top5}"
    ),
    "AU CoR - Fatigue Management Inspection": (
        "The CoR Fatigue Management Inspection verifies driver work-rest records and fatigue "
        "management plans for heavy vehicle operators supporting {task}, reducing impaired-driving "
        "incidents. {top5}"
    ),
    "AU ERSED Inspection": (
        "The ERSED Inspection verifies erosion, sediment, and drainage controls are functional "
        "and adequately maintained to prevent sediment runoff to waterways during {task}. {top5}"
    ),
    "AU - Safety Audit": (
        "The Safety Audit provides management-level verification that all HSE controls for {task} "
        "are implemented, maintained, and effective at the task completion milestone. {top5}"
    ),
    "AU - Environment Audit": (
        "The Environment Audit verifies that all SER environmental controls for {task} are "
        "implemented, maintained, and effective at the task completion milestone. {top5}"
    ),
}

_DEFAULT_REASONING = (
    "This {module} {type_short} verifies that controls applicable to {task} are in place and "
    "effective, reducing the likelihood of an uncontrolled incident. {top5}"
)

_TOP5_SUFFIX = (
    "As the primary fatal/severe risk control for this task, non-implementation presents the "
    "highest residual risk for this horizon and it is therefore flagged as a Top 5 priority."
)
_NOT_TOP5_SUFFIX = (
    "This control supports the layered risk management framework for this task but is not the "
    "primary residual risk control within its horizon."
)


def _risk_detail(task: str, category: str) -> str:
    tmpl = _RISK_DETAILS.get(category, _DEFAULT_RISK_DETAIL)
    short = task[:80] + ("…" if len(task) > 80 else "")
    return tmpl.format(task=short)


def _reasoning(task: str, checklist: str, is_top5: bool, module: str, ctrl_type: str) -> str:
    top5_txt = _TOP5_SUFFIX if is_top5 else _NOT_TOP5_SUFFIX
    short = task[:80] + ("…" if len(task) > 80 else "")
    tmpl = _REASONING.get(checklist, _DEFAULT_REASONING)
    type_short = ctrl_type.split("(")[0].strip() if "(" in ctrl_type else ctrl_type
    return tmpl.format(task=short, top5=top5_txt, module=module, type_short=type_short)


# ── Horizon + due date logic ───────────────────────────────────────────────────

_DUE_OFFSETS: dict[str, int] = {
    "Toolbox Talks": 0,        # same day as task start
    "AU Fatal and Severe Risk (FSR)": 2,   # during activity
    "AU Severe Environmental Risk (SER)": 2,
    "AU Health and Safety Inspections - General": 2,
    "AU Environment Inspection - General": 2,
    "AU Environment Inspection - Assurance": 2,
    "AU Health, Safety & Environment Inspection": 14,  # milestone
    "AU - Safety Audit": 14,
    "AU - Environment Audit": 14,
    "AU - Operational Readiness Review": 14,
}


def _due_date(task_start: datetime, horizon_end: datetime, ctrl_type: str) -> datetime:
    offset = _DUE_OFFSETS.get(ctrl_type, 2)
    due = task_start + timedelta(days=offset)
    return min(due, horizon_end)


def _horizon(task_start: datetime, plan_start: datetime) -> str:
    delta = (task_start - plan_start).days
    if delta <= 30:
        return "30D"
    if delta <= 60:
        return "60D"
    return "90D"


def _horizon_end(plan_start: datetime, horizon: str) -> datetime:
    days = {"30D": 30, "60D": 60, "90D": 90}[horizon]
    return plan_start + timedelta(days=days)


# ── Classification of severity ─────────────────────────────────────────────────

_HIGH_CATS = {
    "Cranes and Lifting", "Prevention of Falls & Dropped Objects",
    "Work In, Over or Adjacent to Water", "Excavation & Ground Penetration",
    "Demolition", "Formwork & Falsework", "Piling", "Scaffolding",
    "General Electrical Safety", "Confined Space", "Hazardous Energy",
    "Working in Rail Environments", "Temporary Waterway Structures",
}
_MEDIUM_CATS = {
    "Traffic Management", "Heavy Vehicles", "Plant and Equipment",
    "Precast Concrete", "Respirable Crystalline Silica", "Hazardous Substances",
    "Dangerous Goods and Chemical Management", "Noise and Vibration",
    "Erosion and Sediment Control", "Water Management", "Spoil, Waste, Resource and Recovery",
}


def _classification(category: str) -> str:
    if category in _HIGH_CATS:
        return "High"
    if category in _MEDIUM_CATS:
        return "Medium"
    return "Low"


# ── Top 5 assignment ───────────────────────────────────────────────────────────


def assign_top5(rows: list[dict]) -> list[dict]:
    """Assign exactly 5 True rows per horizon.

    The True row for each task is its primary FSR/SER inspection row.
    Ranking: (1) Classification, (2) proximity, (3) control count per task.
    """
    rows = copy.deepcopy(rows)

    horizons = ["30D", "60D", "90D"]
    for hz in horizons:
        hz_rows = [r for r in rows if r["Timing"] == hz]
        if not hz_rows:
            continue

        # Group by task
        tasks: dict[str, list[dict]] = {}
        for r in hz_rows:
            tasks.setdefault(r["Job Task.Record No."], []).append(r)

        # Score each task
        def task_score(task_rows: list[dict]) -> tuple:
            cls_map = {"High": 3, "Medium": 2, "Low": 1}
            max_cls = max(cls_map.get(r["Classification"], 1) for r in task_rows)
            start = task_rows[0].get("_start_date", datetime.max)
            return (-max_cls, start, -len(task_rows))

        ranked_tasks = sorted(tasks.items(), key=lambda x: task_score(x[1]))
        top_tasks = [name for name, _ in ranked_tasks[:5]]

        for r in rows:
            if r["Timing"] != hz:
                continue
            task_name = r["Job Task.Record No."]
            is_candidate = task_name in top_tasks
            # The True row = primary FSR/SER for this task
            ctrl_type = r.get("Type", "")
            is_fsr_ser = "FSR" in r.get("Title", "") or "SER" in r.get("Title", "") or \
                         "AU Fatal" in ctrl_type or "AU Severe Env" in ctrl_type
            r["Top 5 Risk"] = "True" if (is_candidate and is_fsr_ser and
                                          r.get("_is_primary_fsr", False)) else "False"

        # Ensure exactly 5 True per horizon
        true_rows = [r for r in rows if r["Timing"] == hz and r["Top 5 Risk"] == "True"]
        if len(true_rows) < 5:
            # Promote next-best rows
            false_rows = sorted(
                [r for r in rows if r["Timing"] == hz and r["Top 5 Risk"] == "False"],
                key=lambda r: ({"High": 0, "Medium": 1, "Low": 2}.get(r["Classification"], 2),
                               r.get("_start_date", datetime.max)),
            )
            needed = 5 - len(true_rows)
            promoted = set()
            for r in false_rows:
                tn = r["Job Task.Record No."]
                if tn not in promoted and len(promoted) < needed:
                    r["Top 5 Risk"] = "True"
                    promoted.add(tn)
        elif len(true_rows) > 5:
            # Demote excess — keep first 5
            count = 0
            for r in rows:
                if r["Timing"] == hz and r["Top 5 Risk"] == "True":
                    count += 1
                    if count > 5:
                        r["Top 5 Risk"] = "False"

    return rows


# ── Main plan builder ──────────────────────────────────────────────────────────


def build_risk_plan(
    activities: list[dict],
    plan_start: datetime,
    person_responsible: str = "[Assignee]",
) -> list[dict]:
    """Build the full list of risk plan rows from schedule activities."""
    rows: list[dict] = []

    for act in activities:
        desc = act["description"]
        wbs = act.get("wbs", "")
        task_start = act["start_date"]
        hz = _horizon(task_start, plan_start)
        hz_end = _horizon_end(plan_start, hz)

        categories = classify_activity(desc, wbs)
        primary = primary_category(categories)
        controls = controls_for_task(categories, primary, desc)

        # Determine which control is the primary FSR for this task
        primary_fsr_checklist = CATEGORY_TO_FSR.get(primary, "")

        for i, ctrl in enumerate(controls):
            ctrl_cat = ctrl["category"]
            ctrl_checklist = ctrl["checklist"]
            due = _due_date(task_start, hz_end, ctrl["type"])

            rows.append(
                {
                    "Timing": hz,
                    "Job Task.Record No.": desc,
                    "Top 5 Risk": "False",
                    "Risk Category.Name": ctrl_cat,
                    "Classification": _classification(ctrl_cat),
                    "Risk Details": _risk_detail(desc, ctrl_cat),
                    "Module": ctrl["module"],
                    "Type": ctrl["type"],
                    "Title": ctrl_checklist,
                    "Person Responsible": person_responsible,
                    "Due Date": due.strftime("%d/%m/%Y"),
                    "Reasoning": _reasoning(desc, ctrl_checklist, False, ctrl["module"], ctrl["type"]),
                    "_start_date": task_start,
                    "_is_primary_fsr": ctrl_checklist == primary_fsr_checklist,
                }
            )

    rows = assign_top5(rows)

    # Sort: 30D → 60D → 90D, tasks contiguous
    order = {"30D": 0, "60D": 1, "90D": 2}
    rows.sort(key=lambda r: (order[r["Timing"]], r.get("_start_date", datetime.min), r["Job Task.Record No."]))

    # Update Reasoning for Top 5 rows
    for r in rows:
        if r["Top 5 Risk"] == "True":
            r["Reasoning"] = _reasoning(
                r["Job Task.Record No."], r["Title"], True, r["Module"], r["Type"]
            )

    return rows


# ── Excel writing ──────────────────────────────────────────────────────────────

COLUMNS = [
    "Timing", "Job Task.Record No.", "Top 5 Risk", "Risk Category.Name",
    "Classification", "Risk Details", "Module", "Type", "Title",
    "Person Responsible", "Due Date", "Reasoning",
]


def write_excel(
    project_name: str,
    work_pack: str,
    start_month_year: str,
    rows: list[dict],
) -> bytes:
    """Populate the base template and return bytes of the completed workbook."""
    template_path = DATA_DIR / "base_template.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Template"]

    # B1 = Project/Office/Depot, B2 = plan name, leave B3 (=TODAY()) untouched
    ws["B1"] = project_name
    ws["B2"] = f"30/60/90 - {work_pack} - {start_month_year}"

    # Write data from row 6 onwards — values only
    for i, row in enumerate(rows):
        r = 6 + i
        for j, col in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=j).value = row.get(col, "")

    # Trim unused pre-styled rows below data
    last_data_row = 5 + len(rows)
    max_row = ws.max_row
    if max_row > last_data_row:
        for del_row in range(max_row, last_data_row, -1):
            ws.delete_rows(del_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
